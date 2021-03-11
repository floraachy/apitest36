"""
======================================
Author: Flora.Chen
Time: 2021/3/8 13:45
~ _ ~ ~ _ ~ ~ _ ~ ~ _ ~ ~ _ ~ ~ _ ~ 
======================================
"""
import openpyxl

class HandleExcel:
    """
    从excel读取写入数据的类
    """
    def __init__(self, file):
        """
        初始化用例文件
        :param file: 文件绝对路径
        """
        self.file = file

    def read(self, sheet_name):
        """
        读取excel数据并返回
        :param sheet_name: 表单名称
        :return: 存在传入的表单, 返回表单数据，不存在则返回空
        """
        # 创建一个工作簿工作对象
        workbook = openpyxl.open(self.file)
        # 跟上面那句一个意思 workbook = openpyxl.load_workbook(self.file)

        # 获取excel当中所有的sheet，返回的是一个列表
        sheets = workbook.sheetnames
        # 获取sheet对象
        if sheet_name in sheets:
            sheet = workbook[sheet_name]
            all_values = list(sheet.values)
            header = all_values[0]

            data = []
            for i in all_values[1:]:
                data.append(dict(zip(header, i)))
            # 关闭excel
            workbook.close()
            return data
        else:
            # 关闭excel
            workbook.close()
            return None
    def write(self, sheet_name, row, column, data):
        """
        往excel写入数据
        :param sheet_name: 表单名称
        :param row: 要写入的行
        :param column: 要写入的列
        :param data: 要写入的数据
        :return: None
        """
        workbook = openpyxl.open(self.file)
        sheets = workbook.sheetnames
        if sheet_name in sheets:
            sheet = workbook[sheet_name]
            sheet.cell(row=row, column=column, value=data)
            # 更上面写法效果一样 sheet.cell(row=row, column=column).value = data

            # 保存并关闭文件
            workbook.save(self.file)
            workbook.close()
        else:
            # 关闭文件
            workbook.close()
            print("{}不存在，请检查~".format(sheet_name))



if __name__ == "__main__":
    excel = HandleExcel("D:\PythonProject\TestDeveloperPlatform36\data\cases.xlsx")
    print(excel.read("register"))
    excel.write(sheet_name="register", row=2, column=9, data="test1")



